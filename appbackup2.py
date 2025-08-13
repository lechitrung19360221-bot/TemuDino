#!/usr/bin/env python3
"""
Advanced Mockup Tool with Real-time Preview
Supports design overlays, watermarks, and batch processing
"""

import sys
import os
from pathlib import Path
from typing import Optional, List, Dict, Any
import glob
import json
import tempfile
import requests
from PIL import Image, ImageEnhance
import math
import re
import random
import string
import numpy as np
from urllib.parse import urlsplit

# Add project root to path for imports
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, 
    QPushButton, QSpinBox, QDoubleSpinBox, QComboBox, QCheckBox,
    QLineEdit, QFileDialog, QGraphicsView, QGraphicsScene,
    QGraphicsPixmapItem, QScrollArea, QGroupBox, QFrame, QMessageBox, QProgressBar,
    QSizePolicy
)
from PySide6.QtCore import Qt, QSize, QPointF, QRectF, QEvent, QSettings
from PySide6.QtGui import QPixmap, QPainter, QWheelEvent, QMouseEvent, QKeyEvent, QFont, QIcon, QGuiApplication, QImage

# Import CLI functions
from mockup_cli.cli import render_one_rect_pixels as render_simple, upload_to_imgbb

class BlendPixmapItem(QGraphicsPixmapItem):
    """Custom pixmap item with blend mode support"""
    
    def __init__(self, pixmap: QPixmap, blend_mode: str = "normal"):
        super().__init__(pixmap)
        self.blend_mode = blend_mode
        self.setFlag(QGraphicsPixmapItem.ItemIsMovable, True)
        self.setFlag(QGraphicsPixmapItem.ItemIsSelectable, True)
        self.setAcceptHoverEvents(True)
    
    def set_blend_mode(self, mode: str):
        self.blend_mode = mode
        self.update()
    
    def paint(self, painter: QPainter, option, widget):
        if self.blend_mode == "normal":
            super().paint(painter, option, widget)
        else:
            # Apply blend mode effect
            painter.setCompositionMode(self._get_composition_mode())
            super().paint(painter, option, widget)
    
    def _get_composition_mode(self):
        blend_map = {
            "multiply": QPainter.CompositionMode_Multiply,
            "screen": QPainter.CompositionMode_Screen,
            "overlay": QPainter.CompositionMode_Overlay,
            "lighten": QPainter.CompositionMode_Lighten,
            "darken": QPainter.CompositionMode_Darken
        }
        return blend_map.get(self.blend_mode, QPainter.CompositionMode_SourceOver)

class MockupApp(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("DINOTOOL TEMU GENERATOR MOCKUP by TRUNGLEE")
        self.setMinimumSize(1200, 800)
        
        # File paths
        self.mockup_path: Optional[Path] = None
        self.design_path: Optional[Path] = None
        self.watermark_path: Optional[Path] = None
        self.preview_img: Optional[Image.Image] = None
        self.mockup_size: Optional[tuple[int, int]] = None
        
        # Graphics scene
        self.scene = QGraphicsScene(self)
        self.view = QGraphicsView(self.scene, self)
        self.view.setRenderHints(QPainter.Antialiasing | QPainter.SmoothPixmapTransform)
        self.view.setDragMode(QGraphicsView.ScrollHandDrag)
        self.view.setMinimumSize(600, 400)
        
        # Graphics items
        self.bg_item: Optional[QGraphicsPixmapItem] = None
        self.overlay_item: Optional[BlendPixmapItem] = None  # main design
        self.wm_item: Optional[BlendPixmapItem] = None  # watermark
        self.active_item: Optional[BlendPixmapItem] = None
        self.overlay_original_size: Optional[tuple[int, int]] = None
        self.wm_original_size: Optional[tuple[int, int]] = None
        # internal state
        self._is_scaling: bool = False
        
        # Lock states (no global lock)
        self.design_locked = False
        self.watermark_locked = False
        self.stop_requested = False
        
        # UI setup
        self._setup_ui()
        self._setup_connections()
        self._setup_interactions()
        # QSS support: only load once at startup (no in-app editing)
        
        # Set initial active item
        self.active_item = self.overlay_item
        # Ensure transform controls reflect current locks/active layer
        self._update_transform_controls_enabled()
        # Load persisted settings
        try:
            self._load_settings()
        except Exception:
            pass
        
    def _setup_ui(self):
        """Setup the user interface"""
        main_layout = QVBoxLayout()

        # App header with optional logo
        self._add_header(main_layout)
        
        # Create scroll area for controls
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # File Selection Group
        file_group = QGroupBox("FILE SELECTION")
        file_group.setObjectName("file_group")
        self._set_group_title_font(file_group, 64)
        file_layout = QVBoxLayout(file_group)
        file_layout.setSpacing(6)
        file_layout.setContentsMargins(10, 8, 10, 10)
        
        # Mockup selection
        mockup_row = QHBoxLayout()
        self.mockup_label = QLabel("Mockup: (chưa chọn)")
        self.mockup_label.setObjectName("file_info_label")
        self.pick_mockup_btn = QPushButton("Chọn Mockup JPG/PNG")
        self.pick_mockup_btn.setObjectName("btn_pick_mockup")
        self.pick_mockup_btn.clicked.connect(self.pick_mockup)
        mockup_row.addWidget(self.mockup_label)
        mockup_row.addWidget(self.pick_mockup_btn)
        file_layout.addLayout(mockup_row)
        
        # Design selection
        design_row = QHBoxLayout()
        self.design_label = QLabel("Design: (chưa chọn)")
        self.design_label.setObjectName("file_info_label")
        self.pick_design_btn = QPushButton("Chọn Design PNG")
        self.pick_design_btn.setObjectName("btn_pick_design")
        self.pick_design_btn.clicked.connect(self.pick_design)
        design_row.addWidget(self.design_label)
        design_row.addWidget(self.pick_design_btn)
        file_layout.addLayout(design_row)
        
        # Watermark selection
        watermark_row = QHBoxLayout()
        self.watermark_label = QLabel("Watermark: (không bắt buộc)")
        self.watermark_label.setObjectName("file_info_label")
        self.pick_watermark_btn = QPushButton("Chọn Watermark PNG")
        self.pick_watermark_btn.setObjectName("btn_pick_watermark")
        self.pick_watermark_btn.clicked.connect(self.pick_watermark)
        watermark_row.addWidget(self.watermark_label)
        watermark_row.addWidget(self.pick_watermark_btn)
        file_layout.addLayout(watermark_row)
        
        scroll_layout.addWidget(file_group)
        
        # Transform Controls Group
        transform_group = QGroupBox("TRANSFORM CONTROLS")
        transform_group.setObjectName("transform_group")
        self._set_group_title_font(transform_group, 64)
        transform_layout = QVBoxLayout(transform_group)
        transform_layout.setSpacing(6)
        transform_layout.setContentsMargins(10, 8, 10, 10)
        
        # Align controls neatly using a grid (3 rows x 6 columns)
        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(10)

        # Row 0
        grid.addWidget(QLabel("Anchor:"), 0, 0)
        self.anchor_combo = QComboBox(); self.anchor_combo.addItems(["center", "topleft"]) 
        grid.addWidget(self.anchor_combo, 0, 1)
        grid.addWidget(QLabel("X:"), 0, 2)
        self.x_spin = QSpinBox(); self.x_spin.setRange(0, 100000); self.x_spin.setValue(1000)
        grid.addWidget(self.x_spin, 0, 3)
        grid.addWidget(QLabel("Y:"), 0, 4)
        self.y_spin = QSpinBox(); self.y_spin.setRange(0, 100000); self.y_spin.setValue(800)
        grid.addWidget(self.y_spin, 0, 5)

        # Row 1
        grid.addWidget(QLabel("Width:"), 1, 0)
        self.w_spin = QSpinBox(); self.w_spin.setRange(1, 100000); self.w_spin.setValue(800)
        grid.addWidget(self.w_spin, 1, 1)
        grid.addWidget(QLabel("Height:"), 1, 2)
        self.h_spin = QSpinBox(); self.h_spin.setRange(1, 100000); self.h_spin.setValue(600)
        grid.addWidget(self.h_spin, 1, 3)
        grid.addWidget(QLabel("Rotation:"), 1, 4)
        self.rot_spin = QDoubleSpinBox(); self.rot_spin.setRange(-360.0, 360.0); self.rot_spin.setDecimals(2)
        grid.addWidget(self.rot_spin, 1, 5)

        # Row 2
        grid.addWidget(QLabel("Opacity:"), 2, 0)
        self.opacity_spin = QDoubleSpinBox(); self.opacity_spin.setRange(0.0, 1.0); self.opacity_spin.setSingleStep(0.05); self.opacity_spin.setValue(1.0)
        grid.addWidget(self.opacity_spin, 2, 1)
        grid.addWidget(QLabel("Aspect:"), 2, 2)
        self.aspect_combo = QComboBox(); self.aspect_combo.addItems(["contain", "cover"]) 
        grid.addWidget(self.aspect_combo, 2, 3)
        grid.addWidget(QLabel("Blend:"), 2, 4)
        self.blend_combo = QComboBox(); self.blend_combo.addItems(["normal", "multiply", "screen", "overlay", "lighten", "darken"]) 
        grid.addWidget(self.blend_combo, 2, 5)

        # Stretch input columns so edges align
        for c in (1, 3, 5):
            grid.setColumnStretch(c, 1)

        transform_layout.addLayout(grid)
        
        scroll_layout.addWidget(transform_group)
        
        # Normalize sizes and alignment for transform controls
        self._uniform_transform_controls()
        
        # Layer Control Group
        layer_group = QGroupBox("LAYER CONTROL")
        layer_group.setObjectName("layer_group")
        self._set_group_title_font(layer_group, 64)
        layer_layout = QVBoxLayout(layer_group)
        layer_layout.setSpacing(6)
        layer_layout.setContentsMargins(10, 8, 10, 10)
        
        # Layer selection buttons
        layer_buttons = QHBoxLayout()
        self.focus_design_btn = QPushButton("Chỉnh DESIGN")
        self.focus_watermark_btn = QPushButton("Chỉnh WATERMARK")
        self.focus_design_btn.clicked.connect(lambda: self.set_active_layer("design"))
        self.focus_watermark_btn.clicked.connect(lambda: self.set_active_layer("watermark"))
        layer_buttons.addWidget(self.focus_design_btn)
        layer_buttons.addWidget(self.focus_watermark_btn)
        layer_layout.addLayout(layer_buttons)
        
        # Active layer indicator
        self.active_label = QLabel("Active: DESIGN")
        self.active_label.setObjectName("active_label")
        layer_layout.addWidget(self.active_label)
        
        # Lock checkboxes
        lock_row = QHBoxLayout()
        self.lock_design_chk = QCheckBox("Khóa DESIGN")
        self.lock_watermark_chk = QCheckBox("Khóa WATERMARK")
        self.lock_design_chk.toggled.connect(self.toggle_design_lock)
        self.lock_watermark_chk.toggled.connect(self.toggle_watermark_lock)
        lock_row.addWidget(self.lock_design_chk)
        lock_row.addWidget(self.lock_watermark_chk)
        layer_layout.addLayout(lock_row)
        
        # Add preview view here (between Preview Controls and Batch Export)
        preview_view_group = QGroupBox("PREVIEW")
        preview_view_group.setObjectName("preview_view_group")
        self._set_group_title_font(preview_view_group, 64)
        preview_view_layout = QVBoxLayout(preview_view_group)
        preview_view_layout.setSpacing(6)
        preview_view_layout.setContentsMargins(8, 8, 8, 8)
        
        # Controls row (buttons) on top of the design frame
        controls_row = QHBoxLayout()
        self.preview_btn = QPushButton("Preview")
        self.preview_btn.setObjectName("btn_preview")
        self.preview_btn.clicked.connect(self.on_preview)
        self.reset_overlay_btn = QPushButton("Reset overlay")
        self.reset_overlay_btn.setObjectName("btn_reset_overlay")
        self.reset_overlay_btn.clicked.connect(self.reset_overlay)
        self.center_overlay_btn = QPushButton("Center overlay")
        self.center_overlay_btn.setObjectName("btn_center_overlay")
        self.center_overlay_btn.clicked.connect(self.center_overlay)
        for b in (self.preview_btn, self.reset_overlay_btn, self.center_overlay_btn):
            b.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        controls_row.addWidget(self.preview_btn)
        controls_row.addWidget(self.reset_overlay_btn)
        controls_row.addWidget(self.center_overlay_btn)
        preview_view_layout.addLayout(controls_row)

        # The design preview frame
        preview_view_layout.addWidget(self.view)
        scroll_layout.addWidget(preview_view_group)

        # Move Layer Control BELOW the Preview frame
        scroll_layout.addWidget(layer_group)

        # (In-app theme editor removed by request)
        
        # Batch Export Group
        export_group = QGroupBox("BATCH EXPORT")
        export_group.setObjectName("export_group")
        self._set_group_title_font(export_group, 64)
        export_layout = QVBoxLayout(export_group)
        export_layout.setSpacing(6)
        export_layout.setContentsMargins(10, 8, 10, 10)
        
        # Nguồn thiết kế (một hàng) — đưa lên đầu
        mode_row = QHBoxLayout()
        mode_row.addWidget(self._create_form_label("Nguồn thiết kế:"))
        self.source_folder_chk = QCheckBox("Folder")
        self.source_json_chk = QCheckBox("JSON")
        mode_row.addWidget(self.source_folder_chk)
        mode_row.addWidget(self.source_json_chk)
        mode_row.addStretch(1)
        export_layout.addLayout(mode_row)

        # Output directory
        out_row = QHBoxLayout()
        out_row.addWidget(self._create_form_label("Output Directory:"))
        self.out_dir_edit = QLineEdit()
        self.out_dir_edit.setPlaceholderText("outputs")
        self.out_dir_btn = QPushButton("Chọn thư mục")
        self.out_dir_btn.clicked.connect(self.pick_out_dir)
        self._style_select_button(self.out_dir_btn)
        self._style_path_edit(self.out_dir_edit)
        out_row.addWidget(self.out_dir_edit, 1)
        out_row.addWidget(self.out_dir_btn)
        export_layout.addLayout(out_row)
        
        self.source_folder_chk.setChecked(True)
        self.source_json_chk.setChecked(False)
        # Click -> behave like radio buttons
        self.source_folder_chk.clicked.connect(self.select_folder_mode)
        self.source_json_chk.clicked.connect(self.select_json_mode)

        # Design sources (folder)
        designs_row = QHBoxLayout()
        designs_row.addWidget(self._create_form_label("Folder/Pattern:"))
        self.designs_dir_edit = QLineEdit()
        self.designs_dir_edit.setPlaceholderText("designs/*.png hoặc chọn thư mục")
        self.designs_dir_btn = QPushButton("Chọn thư mục")
        self.designs_dir_btn.clicked.connect(self.pick_designs_dir)
        self._style_select_button(self.designs_dir_btn)
        self._style_path_edit(self.designs_dir_edit)
        designs_row.addWidget(self.designs_dir_edit, 1)
        designs_row.addWidget(self.designs_dir_btn)
        export_layout.addLayout(designs_row)
        
        # JSON URL support (mutually exclusive with folder source)
        # JSON inputs split into two rows for clarity/alignment
        json_path_row = QHBoxLayout()
        json_path_row.addWidget(self._create_form_label("Chọn đường dẫn JSON:"))
        self.json_path_edit = QLineEdit()
        self.json_path_edit.setPlaceholderText("Đường dẫn file JSON")
        self.json_path_edit.setReadOnly(True)
        self._style_path_edit(self.json_path_edit)
        self.pick_json_btn = QPushButton("Chọn file JSON")
        self.pick_json_btn.setObjectName("btn_pick_json")
        self.pick_json_btn.clicked.connect(self.pick_json_file)
        self._style_select_button(self.pick_json_btn)
        json_path_row.addWidget(self.json_path_edit, 1)
        json_path_row.addWidget(self.pick_json_btn)
        export_layout.addLayout(json_path_row)

        temp_dir_row = QHBoxLayout()
        temp_dir_row.addWidget(self._create_form_label("Chọn thư mục tạm:"))
        self.temp_dir_edit = QLineEdit()
        self.temp_dir_edit.setPlaceholderText("designs_temp")
        self._style_path_edit(self.temp_dir_edit)
        self.temp_dir_btn = QPushButton("Chọn thư mục tạm")
        self.temp_dir_btn.setObjectName("btn_pick_temp_dir")
        self.temp_dir_btn.clicked.connect(self.pick_temp_dir)
        self._style_select_button(self.temp_dir_btn)
        temp_dir_row.addWidget(self.temp_dir_edit, 1)
        temp_dir_row.addWidget(self.temp_dir_btn)
        export_layout.addLayout(temp_dir_row)
        
        # Output pattern
        pattern_row = QHBoxLayout()
        pattern_row.addWidget(self._create_form_label("Output Pattern:"))
        self.pattern_edit = QLineEdit("{name}_mockup.png")
        self._style_path_edit(self.pattern_edit)
        pattern_row.addWidget(self.pattern_edit, 1)
        export_layout.addLayout(pattern_row)
        
        # ImgBB upload
        upload_row = QHBoxLayout()
        self.upload_chk = QCheckBox("Upload to ImgBB")
        try:
            self.upload_chk.setMinimumWidth(200)
        except Exception:
            pass
        self.imgbb_key_edit = QLineEdit()
        self._style_path_edit(self.imgbb_key_edit)
        self.imgbb_key_edit.setPlaceholderText("IMGBB API KEY (hoặc để trống dùng env IMGBB_API_KEY)")
        upload_row.addWidget(self.upload_chk)
        upload_row.addWidget(self.imgbb_key_edit, 1)
        export_layout.addLayout(upload_row)
        
        # Excel controls
        excel_row1 = QHBoxLayout()
        self.export_excel_chk = QCheckBox("Ghi Excel (Temu)")
        try:
            self.export_excel_chk.setMinimumWidth(200)
        except Exception:
            pass
        excel_row1.addWidget(self.export_excel_chk)
        self.excel_template_edit = QLineEdit()
        self._style_path_edit(self.excel_template_edit)
        self.excel_template_btn = QPushButton("Chọn template")
        self.excel_template_btn.setObjectName("btn_excel_choose_template")
        self.excel_template_btn.clicked.connect(self.pick_excel_template)
        self._style_select_button(self.excel_template_btn)
        excel_row1.addWidget(self.excel_template_edit, 1)
        excel_row1.addWidget(self.excel_template_btn)
        export_layout.addLayout(excel_row1)

        excel_row2 = QHBoxLayout()
        excel_row2.addWidget(self._create_form_label("Excel Output Folder:"))
        self.excel_out_edit = QLineEdit()
        self._style_path_edit(self.excel_out_edit)
        self.excel_out_btn = QPushButton("Chọn thư mục")
        self.excel_out_btn.setObjectName("btn_excel_out_dir")
        self.excel_out_btn.clicked.connect(self.pick_excel_out)
        self._style_select_button(self.excel_out_btn)
        excel_row2.addWidget(self.excel_out_edit, 1)
        excel_row2.addWidget(self.excel_out_btn)
        export_layout.addLayout(excel_row2)
        # Add vertical space before the action buttons
        try:
            export_layout.addSpacing(16)
        except Exception:
            pass

        # Export button
        self.export_btn = QPushButton("Export hàng loạt")
        self.export_btn.setObjectName("export_btn")
        self.export_btn.clicked.connect(self.on_export_batch)
        # Stop button
        self.stop_btn = QPushButton("Dừng")
        self.stop_btn.setObjectName("stop_btn")
        self.stop_btn.clicked.connect(self.on_stop_clicked)
        # Reset-all button
        self.reset_all_btn = QPushButton("Reset toàn bộ")
        self.reset_all_btn.setObjectName("reset_all_btn")
        self.reset_all_btn.clicked.connect(self.reset_all)
        row_actions = QHBoxLayout()
        row_actions.setSpacing(12)
        try:
            for b in (self.export_btn, self.stop_btn, self.reset_all_btn):
                b.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                b.setMinimumHeight(44)
        except Exception:
            pass
        row_actions.addWidget(self.export_btn, 1)
        row_actions.addWidget(self.stop_btn, 1)
        row_actions.addWidget(self.reset_all_btn, 1)
        export_layout.addLayout(row_actions)

        # Initialize source mode UI state
        self.update_source_mode()
        
        scroll_layout.addWidget(export_group)
        
        # Set scroll area
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        # Không giới hạn chiều cao để luôn cuộn xuống được đến nút Export
        # scroll_area.setMaximumHeight(600)
        
        main_layout.addWidget(scroll_area)

        # Bottom full-width progress bar (green)
        bottom_row = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("%p%")
        self.progress_bar.setAlignment(Qt.AlignCenter)
        self.progress_bar.setRange(0, 1)
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(16)
        # Style comes from global QSS
        # Hidden status label placeholder for code paths using it
        self.status_label = QLabel("")
        bottom_row.addWidget(self.progress_bar, 1)
        main_layout.addLayout(bottom_row)
        
        self.setLayout(main_layout)

    def _style_select_button(self, button: QPushButton) -> None:
        """Make selection buttons the same size and aligned visually to the right."""
        try:
            button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            button.setFixedWidth(160)
            button.setFixedHeight(36)
        except Exception:
            pass

    def _style_path_edit(self, edit: QLineEdit) -> None:
        try:
            edit.setFixedHeight(36)
        except Exception:
            pass

    def _create_form_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        try:
            lbl.setMinimumWidth(200)
        except Exception:
            pass
        return lbl

    def _add_header(self, parent_layout: QVBoxLayout) -> None:
        """Add a simple header with an optional logo at the very top of the app.

        Place your logo at mockup_gui/assets/logo.png and it will be shown automatically.
        """
        try:
            header_widget = QWidget()
            header_layout = QHBoxLayout(header_widget)
            header_layout.setContentsMargins(8, 8, 8, 8)
            header_layout.setSpacing(10)

            logo_label = QLabel()
            logo_label.setAttribute(Qt.WA_TranslucentBackground, True)
            # Use logo at mockup_gui/logo.png
            logo_path = Path(__file__).resolve().parent / "logo.png"
            if logo_path.exists():
                pix = QPixmap(str(logo_path))
                if not pix.isNull():
                    # Show a small, crisp logo at top-left (scaled smoothly)
                    target_h = 60
                    # Smooth downscale to avoid jagged edges
                    scaled = pix if pix.height() == target_h else pix.scaledToHeight(target_h, Qt.SmoothTransformation)
                    logo_label.setPixmap(scaled)
                    logo_label.setScaledContents(False)
                    header_widget.setFixedHeight(target_h + 16)
                    try:
                        QApplication.instance().setWindowIcon(QIcon(str(logo_path)))
                    except Exception:
                        pass

            header_layout.addWidget(logo_label)
            header_layout.addStretch(1)

            parent_layout.addWidget(header_widget)
        except Exception:
            pass

    def _set_group_title_font(self, group: QGroupBox, point_size: int) -> None:
        f = group.font()
        f.setPointSize(int(point_size))
        # Use the heaviest available weight
        try:
            f.setWeight(QFont.Black)
        except Exception:
            f.setBold(True)
        group.setFont(f)
        
    def _setup_connections(self):
        """Setup signal connections"""
        # Connect spin boxes to update functions
        self.x_spin.valueChanged.connect(self.update_active_item_from_spins)
        self.y_spin.valueChanged.connect(self.update_active_item_from_spins)
        self.w_spin.valueChanged.connect(self.update_active_item_from_spins)
        self.h_spin.valueChanged.connect(self.update_active_item_from_spins)
        self.rot_spin.valueChanged.connect(self.update_active_item_from_spins)
        self.opacity_spin.valueChanged.connect(self.update_active_item_from_spins)
        self.blend_combo.currentTextChanged.connect(self.update_active_item_from_spins)
        
        # Connect anchor combo
        self.anchor_combo.currentTextChanged.connect(self.update_active_item_from_spins)
        self.aspect_combo.currentTextChanged.connect(self.update_active_item_from_spins)
        
    def _set_transform_controls_enabled(self, enabled: bool) -> None:
        widgets = [
            self.x_spin, self.y_spin, self.w_spin, self.h_spin,
            self.rot_spin, self.opacity_spin,
            self.anchor_combo, self.aspect_combo, self.blend_combo,
        ]
        for w in widgets:
            try:
                w.setEnabled(enabled)
            except Exception:
                pass

    def _update_transform_controls_enabled(self) -> None:
        # Disable transforms if the active layer is locked
        if self.active_item == self.overlay_item and self.design_locked:
            self._set_transform_controls_enabled(False)
        elif self.active_item == self.wm_item and self.watermark_locked:
            self._set_transform_controls_enabled(False)
        else:
            self._set_transform_controls_enabled(True)

    def _setup_interactions(self):
        """Setup mouse and keyboard interactions"""
        self.view.viewport().installEventFilter(self)
        self.view.setFocusPolicy(Qt.StrongFocus)

    def _uniform_transform_controls(self) -> None:
        """Make transform inputs consistent in size and alignment for clean layout."""
        try:
            widgets = [
                getattr(self, 'anchor_combo', None),
                getattr(self, 'x_spin', None),
                getattr(self, 'y_spin', None),
                getattr(self, 'w_spin', None),
                getattr(self, 'h_spin', None),
                getattr(self, 'rot_spin', None),
                getattr(self, 'aspect_combo', None),
                getattr(self, 'blend_combo', None),
                getattr(self, 'opacity_spin', None),
            ]
            for w in widgets:
                if w is None:
                    continue
                # Height matches stylesheet target and radius looks smoother
                w.setFixedHeight(28)
                # Give consistent minimum width so columns line up
                w.setMinimumWidth(360)
                # Center numeric values where applicable
                if hasattr(w, 'setAlignment'):
                    try:
                        w.setAlignment(Qt.AlignCenter)
                    except Exception:
                        pass
                # Center combo text nicely without allowing editing
                from PySide6.QtWidgets import QComboBox
                if isinstance(w, QComboBox):
                    try:
                        w.setEditable(True)
                        if w.lineEdit():
                            w.lineEdit().setReadOnly(True)
                            w.lineEdit().setAlignment(Qt.AlignCenter)
                            w.lineEdit().setFrame(False)
                    except Exception:
                        pass
        except Exception:
            pass


    # Settings persistence
    def _load_settings(self) -> None:
        s = QSettings("Temu", "MockupApp")
        # IO
        self.designs_dir_edit.setText(s.value("io/designs_pattern", ""))
        self.out_dir_edit.setText(s.value("io/out_dir", "outputs"))
        self.pattern_edit.setText(s.value("io/pattern", "{name}_mockup.png"))
        # JSON mode
        if hasattr(self, 'source_json_chk'):
            json_mode = s.value("io/json_mode", False, type=bool)
            self.source_json_chk.setChecked(json_mode)
            self.source_folder_chk.setChecked(not json_mode)
            self.update_source_mode()
        if hasattr(self, 'json_path_edit'):
            self.json_path_edit.setText(s.value("io/json_path", ""))
        if hasattr(self, 'temp_dir_edit'):
            self.temp_dir_edit.setText(s.value("io/temp_dir", "designs_temp"))
        # Transforms
        self.anchor_combo.setCurrentText(s.value("transform/anchor", "center"))
        self.aspect_combo.setCurrentText(s.value("transform/aspect", "contain"))
        self.blend_combo.setCurrentText(s.value("transform/blend", "normal"))
        for key, widget, cast in [
            ("transform/x", self.x_spin, int),
            ("transform/y", self.y_spin, int),
            ("transform/w", self.w_spin, int),
            ("transform/h", self.h_spin, int),
            ("transform/rot", self.rot_spin, float),
            ("transform/opacity", self.opacity_spin, float),
        ]:
            val = s.value(key)
            if val is not None:
                try:
                    widget.setValue(cast(val))
                except Exception:
                    pass
        # Locks
        self.lock_design_chk.setChecked(s.value("locks/design", False, type=bool))
        self.lock_watermark_chk.setChecked(s.value("locks/watermark", False, type=bool))
        # Upload
        self.upload_chk.setChecked(s.value("upload/enabled", False, type=bool))
        self.imgbb_key_edit.setText(s.value("upload/key", ""))

    def _save_settings(self) -> None:
        s = QSettings("Temu", "MockupApp")
        s.setValue("io/designs_pattern", self.designs_dir_edit.text())
        s.setValue("io/out_dir", self.out_dir_edit.text())
        s.setValue("io/pattern", self.pattern_edit.text())
        if hasattr(self, 'source_json_chk'):
            s.setValue("io/json_mode", bool(self.source_json_chk.isChecked()))
        if hasattr(self, 'json_path_edit'):
            s.setValue("io/json_path", self.json_path_edit.text())
        if hasattr(self, 'temp_dir_edit'):
            s.setValue("io/temp_dir", self.temp_dir_edit.text())
        s.setValue("transform/anchor", self.anchor_combo.currentText())
        s.setValue("transform/aspect", self.aspect_combo.currentText())
        s.setValue("transform/blend", self.blend_combo.currentText())
        s.setValue("transform/x", self.x_spin.value())
        s.setValue("transform/y", self.y_spin.value())
        s.setValue("transform/w", self.w_spin.value())
        s.setValue("transform/h", self.h_spin.value())
        s.setValue("transform/rot", self.rot_spin.value())
        s.setValue("transform/opacity", self.opacity_spin.value())
        s.setValue("locks/design", bool(self.lock_design_chk.isChecked()))
        s.setValue("locks/watermark", bool(self.lock_watermark_chk.isChecked()))
        s.setValue("upload/enabled", bool(self.upload_chk.isChecked()))
        s.setValue("upload/key", self.imgbb_key_edit.text())

    def closeEvent(self, event) -> None:  # type: ignore[override]
        try:
            self._save_settings()
        finally:
            super().closeEvent(event)

    def update_source_mode(self):
        # Ensure mutual exclusivity between Folder and JSON modes
        if self.source_json_chk.isChecked():
            self.source_folder_chk.setChecked(False)
            self.designs_dir_edit.setEnabled(False)
            self.designs_dir_btn.setEnabled(False)
            self.pick_json_btn.setEnabled(True)
            self.json_path_edit.setEnabled(True)
            self.temp_dir_btn.setEnabled(True)
            self.temp_dir_edit.setEnabled(True)
        elif self.source_folder_chk.isChecked():
            self.source_json_chk.setChecked(False)
            self.designs_dir_edit.setEnabled(True)
            self.designs_dir_btn.setEnabled(True)
            self.pick_json_btn.setEnabled(False)
            self.json_path_edit.setEnabled(False)
            self.temp_dir_btn.setEnabled(False)
            self.temp_dir_edit.setEnabled(False)
        else:
            # Default back to folder mode
            self.source_folder_chk.setChecked(True)
            self.update_source_mode()

    def select_folder_mode(self):
        self.source_folder_chk.setChecked(True)
        self.source_json_chk.setChecked(False)
        self.update_source_mode()

    def select_json_mode(self):
        self.source_json_chk.setChecked(True)
        self.source_folder_chk.setChecked(False)
        self.update_source_mode()
        
    def set_active_layer(self, layer: str):
        """Set the active layer for editing"""
        if layer == "design":
            self.active_item = self.overlay_item
            self.active_label.setText("Active: DESIGN")
            self.focus_design_btn.setStyleSheet("background-color: #007acc; color: white;")
            self.focus_watermark_btn.setStyleSheet("")
        elif layer == "watermark":
            if self.wm_item:
                self.active_item = self.wm_item
                self.active_label.setText("Active: WATERMARK")
                self.focus_watermark_btn.setStyleSheet("background-color: #007acc; color: white;")
                self.focus_design_btn.setStyleSheet("")
            else:
                self.active_label.setText("Active: DESIGN (no watermark)")
                self.active_item = self.overlay_item
        # Rule: đổi layer active KHÔNG thay đổi spin. Chỉ các thao tác chỉnh (spin, wheel, reset/center)
        # mới cập nhật item, còn click không đụng tới thông số.
        # Also refresh enabled state of transform controls
        self._update_transform_controls_enabled()
        return
        
    def toggle_design_lock(self, locked: bool):
        """Toggle design layer lock"""
        self.design_locked = locked
        if self.overlay_item:
            self.overlay_item.setFlag(QGraphicsPixmapItem.ItemIsMovable, not locked)
            self.overlay_item.setFlag(QGraphicsPixmapItem.ItemIsSelectable, not locked)
        # Disable transform controls if active is locked
        self._update_transform_controls_enabled()
        
    def toggle_watermark_lock(self, locked: bool):
        """Toggle watermark layer lock"""
        self.watermark_locked = locked
        if self.wm_item:
            self.wm_item.setFlag(QGraphicsPixmapItem.ItemIsMovable, not locked)
            self.wm_item.setFlag(QGraphicsPixmapItem.ItemIsSelectable, not locked)
        # Disable transform controls if active is locked
        self._update_transform_controls_enabled()
        
    # Global lock removed
    
    def pick_mockup(self):
        """Pick mockup image file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Chọn Mockup", "", "Images (*.jpg *.jpeg *.png *.bmp *.webp)"
        )
        if file_path:
            self.mockup_path = Path(file_path)
            self.mockup_label.setText(f"Mockup: {self.mockup_path.name}")
            self.load_mockup()
    
    def pick_design(self):
        """Pick design image file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Chọn Design", "", "Images (*.png *.jpg *.jpeg *.bmp *.webp)"
        )
        if file_path:
            self.design_path = Path(file_path)
            self.design_label.setText(f"Design: {self.design_path.name}")
            self.load_design()
    
    def pick_watermark(self):
        """Pick watermark image file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Chọn Watermark", "", "Images (*.png *.jpg *.jpeg *.bmp *.webp)"
        )
        if file_path:
            self.watermark_path = Path(file_path)
            self.watermark_label.setText(f"Watermark: {self.watermark_path.name}")
            self.load_watermark()
    
    def pick_out_dir(self):
        """Pick output directory"""
        dir_path = QFileDialog.getExistingDirectory(self, "Chọn thư mục output")
        if dir_path:
            self.out_dir_edit.setText(dir_path)
            # If using JSON temp folder previously, keep as is. No change here.

    def pick_temp_dir(self):
        """Pick temp directory for JSON download"""
        dir_path = QFileDialog.getExistingDirectory(self, "Chọn thư mục tạm để lưu ảnh tải về")
        if dir_path:
            self.temp_dir_edit.setText(dir_path)

    def pick_excel_template(self):
        path, _ = QFileDialog.getOpenFileName(self, "Chọn Excel Template", "", "Excel (*.xlsx)")
        if path:
            self.excel_template_edit.setText(path)

    def pick_excel_out(self):
        dir_path = QFileDialog.getExistingDirectory(self, "Chọn thư mục Excel output")
        if dir_path:
            self.excel_out_edit.setText(dir_path)
    
    def pick_designs_dir(self):
        """Pick designs directory"""
        dir_path = QFileDialog.getExistingDirectory(self, "Chọn thư mục designs")
        if dir_path:
            self.designs_dir_edit.setText(f"{dir_path}/*.png")
            # Mutually exclusive: clear JSON input
            self.json_path_edit.clear()
    
    def pick_json_file(self):
        """Pick JSON file with image URLs"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file JSON", "", "JSON files (*.json)"
        )
        if file_path:
            self.json_path_edit.setText(file_path)
            # Mutually exclusive: clear folder input
            self.designs_dir_edit.clear()
    
    def _resolve_json_items(self, json_path: str) -> list[dict]:
        with open(json_path, 'r') as f:
            data = json.load(f)
        items = []
        if isinstance(data, dict) and "urls" in data:
            for url in data["urls"]:
                items.append({"title": Path(url).stem, "url": url})
        elif isinstance(data, dict) and "items" in data and isinstance(data["items"], list):
            for it in data["items"]:
                if isinstance(it, dict):
                    _url = it.get("url") or it.get("image") or it.get("img") or it.get("link")
                    if _url:
                        items.append({"title": it.get("title") or Path(_url).stem, "url": _url})
        elif isinstance(data, list):
            for it in data:
                if isinstance(it, dict):
                    _url = it.get("url") or it.get("image") or it.get("img") or it.get("link")
                    if _url:
                        items.append({"title": it.get("title") or Path(_url).stem, "url": _url})
                elif isinstance(it, str):
                    items.append({"title": Path(it).stem, "url": it})
        elif isinstance(data, dict):
            for k, v in data.items():
                if isinstance(v, str):
                    items.append({"title": str(k), "url": v})
                elif isinstance(v, dict):
                    _url = v.get("url") or v.get("image") or v.get("img") or v.get("link")
                    if _url:
                        items.append({"title": str(k), "url": _url})
        return items
    
    def load_mockup(self):
        """Load mockup image into scene"""
        if not self.mockup_path:
            return
            
        try:
            # Load with PIL first
            img = Image.open(self.mockup_path)
            self.mockup_size = img.size
            
            # Convert to QPixmap
            img.save("temp_mockup.png")  # Save as PNG for compatibility
            pixmap = QPixmap("temp_mockup.png")
            os.remove("temp_mockup.png")
            
            if pixmap.isNull():
                print("Failed to load mockup image")
                return
            
            # Clear existing background
            if self.bg_item:
                self.scene.removeItem(self.bg_item)
            
            # Add new background
            self.bg_item = QGraphicsPixmapItem(pixmap)
            self.bg_item.setZValue(-1000)  # Background layer
            self.scene.addItem(self.bg_item)
            
            # Fit view to scene
            self.view.fitInView(self.bg_item, Qt.KeepAspectRatio)
            
            # Update scene rect
            self.scene.setSceneRect(self.bg_item.boundingRect())
            
        except Exception as e:
            print(f"Error loading mockup: {e}")
    
    def load_design(self):
        """Load design image into scene"""
        if not self.design_path:
            return
            
        try:
            # Load with PIL first
            img = Image.open(self.design_path)
            
            # Convert to QPixmap
            img.save("temp_design.png")
            pixmap = QPixmap("temp_design.png")
            os.remove("temp_design.png")
            
            if pixmap.isNull():
                print("Failed to load design image")
                return
            
            # Clear existing overlay
            if self.overlay_item:
                self.scene.removeItem(self.overlay_item)
            
            # Create new overlay with blend mode
            blend_mode = self.blend_combo.currentText()
            self.overlay_item = BlendPixmapItem(pixmap, blend_mode)
            self.overlay_item.setZValue(100)  # Above background
            # Rotate/scale around the visual center so export matches preview
            self.overlay_item.setTransformOriginPoint(self.overlay_item.boundingRect().center())
            self.overlay_original_size = (pixmap.width(), pixmap.height())
            
            # Set initial position and size
            self.center_overlay()
            
            # Add to scene
            self.scene.addItem(self.overlay_item)
            
            # Set as active item
            self.active_item = self.overlay_item
            self.update_spins_from_active_item()
            
        except Exception as e:
            print(f"Error loading design: {e}")
    
    def load_watermark(self):
        """Load watermark image into scene"""
        if not self.watermark_path:
            return
            
        try:
            # Load with PIL first
            img = Image.open(self.watermark_path)
            
            # Convert to QPixmap
            img.save("temp_watermark.png")
            pixmap = QPixmap("temp_watermark.png")
            os.remove("temp_watermark.png")
            
            if pixmap.isNull():
                print("Failed to load watermark image")
                return
            
            # Clear existing watermark
            if self.wm_item:
                self.scene.removeItem(self.wm_item)
            
            # Create new watermark with blend mode
            blend_mode = self.blend_combo.currentText()
            self.wm_item = BlendPixmapItem(pixmap, blend_mode)
            self.wm_item.setZValue(200)  # Above design
            self.wm_item.setTransformOriginPoint(self.wm_item.boundingRect().center())
            self.wm_original_size = (pixmap.width(), pixmap.height())
            
            # Set initial position (top-right corner)
            if self.mockup_size:
                wm_rect = self.wm_item.boundingRect()
                x = self.mockup_size[0] - wm_rect.width() - 20
                y = 20
                self.wm_item.setPos(x, y)
            
            # Add to scene
            self.scene.addItem(self.wm_item)
            
        except Exception as e:
            print(f"Error loading watermark: {e}")
    
    def center_overlay(self):
        """Center the active overlay item"""
        if not self.active_item:
            return
        try:
            # Target center is the visual center of background item if available,
            # otherwise the center of the scene rect
            if self.bg_item is not None:
                target_center = self.bg_item.mapToScene(self.bg_item.boundingRect().center())
            else:
                target_center = self.scene.sceneRect().center()

            item_center = self.active_item.mapToScene(self.active_item.boundingRect().center())
            delta = target_center - item_center
            self.active_item.setPos(self.active_item.pos() + delta)
            self.update_spins_from_active_item()
        except Exception:
            pass
    
    def reset_overlay(self):
        """Reset overlay to default position and size"""
        if not self.active_item:
            return
        # Reset size to original image size
        if self.active_item == self.overlay_item and self.overlay_original_size:
            orig_w, orig_h = self.overlay_original_size
        elif self.active_item == self.wm_item and self.wm_original_size:
            orig_w, orig_h = self.wm_original_size
        else:
            rect = self.active_item.boundingRect()
            orig_w, orig_h = int(rect.width()), int(rect.height())

        self.w_spin.setValue(int(orig_w))
        self.h_spin.setValue(int(orig_h))
        self.rot_spin.setValue(0.0)
        self.opacity_spin.setValue(1.0)
        self.anchor_combo.setCurrentText("center")
        self.aspect_combo.setCurrentText("contain")
        self.blend_combo.setCurrentText("normal")

        # Center on mockup and apply
        self.center_overlay()
    
    def update_spins_from_active_item(self):
        """Update spin boxes with current item values"""
        if not self.active_item:
            return
            
        # Block signals to prevent infinite loop
        self.x_spin.blockSignals(True)
        self.y_spin.blockSignals(True)
        self.w_spin.blockSignals(True)
        self.h_spin.blockSignals(True)
        self.rot_spin.blockSignals(True)
        self.opacity_spin.blockSignals(True)
        
        # Current transform and geometry
        transform = self.active_item.transform()
        rect_local = self.active_item.boundingRect()
        center_scene = self.active_item.mapToScene(rect_local.center())
        # Update X/Y depending on anchor policy
        if self.anchor_combo.currentText() == "center":
            self.x_spin.setValue(int(round(center_scene.x())))
            self.y_spin.setValue(int(round(center_scene.y())))
        else:
            # fall back to raw item position (maps local origin to scene)
            pos = self.active_item.pos()
            self.x_spin.setValue(int(round(pos.x())))
            self.y_spin.setValue(int(round(pos.y())))
        # Visual width/height from transform scale
        base_w = rect_local.width()
        base_h = rect_local.height()
        scale_x = float(np.hypot(transform.m11(), transform.m21()))
        scale_y = float(np.hypot(transform.m12(), transform.m22()))
        self.w_spin.setValue(int(round(base_w * (scale_x if scale_x != 0 else 1.0))))
        self.h_spin.setValue(int(round(base_h * (scale_y if scale_y != 0 else 1.0))))
        # Rotation in degrees
        rotation = float(np.degrees(np.arctan2(transform.m21(), transform.m11())))
        self.rot_spin.setValue(rotation)
        
        # Get opacity
        opacity = self.active_item.opacity()
        self.opacity_spin.setValue(opacity)
        
        # Unblock signals
        self.x_spin.blockSignals(False)
        self.y_spin.blockSignals(False)
        self.w_spin.blockSignals(False)
        self.h_spin.blockSignals(False)
        self.rot_spin.blockSignals(False)
        self.opacity_spin.blockSignals(False)
    
    def update_active_item_from_spins(self):
        """Update active item from spin box values"""
        if not self.active_item:
            return
            
        # Get values from spin boxes
        x = self.x_spin.value()
        y = self.y_spin.value()
        w = self.w_spin.value()
        h = self.h_spin.value()
        rotation = self.rot_spin.value()
        opacity = self.opacity_spin.value()
        blend_mode = self.blend_combo.currentText()
        
        # Keep visual center when changing size via spins/wheel
        original_rect = self.active_item.boundingRect()
        self.active_item.setTransformOriginPoint(original_rect.center())
        desired_anchor = self.anchor_combo.currentText()
        # Determine desired target point in scene
        # If user is actively scaling (wheel), force center-anchor behavior
        if getattr(self, "_is_scaling", False) or desired_anchor == "center":
            desired_point_scene = QPointF(float(x), float(y))
        else:
            # 'topleft' uses item position (maps local origin to scene)
            desired_point_scene = QPointF(float(x), float(y))
        # Apply scale+rotation first
        scale_x = max(1e-6, w / original_rect.width())
        scale_y = max(1e-6, h / original_rect.height())
        transform = self.active_item.transform()
        transform.reset()
        transform.scale(scale_x, scale_y)
        transform.rotate(rotation)
        self.active_item.setTransform(transform)
        # Align to desired anchor point precisely
        if desired_anchor == "center":
            current_center_scene = self.active_item.mapToScene(original_rect.center())
            delta = desired_point_scene - current_center_scene
            if delta != QPointF(0, 0):
                self.active_item.setPos(self.active_item.pos() + delta)
        else:
            # For 'topleft', align scene-mapped local origin to desired point
            current_origin_scene = self.active_item.mapToScene(QPointF(0, 0))
            delta = desired_point_scene - current_origin_scene
            if delta != QPointF(0, 0):
                self.active_item.setPos(self.active_item.pos() + delta)
        
        # Apply opacity
        self.active_item.setOpacity(opacity)
        
        # Apply blend mode
        if isinstance(self.active_item, BlendPixmapItem):
            self.active_item.set_blend_mode(blend_mode)
    
    def on_preview(self):
        """Generate preview image that matches the live view exactly"""
        if not self.mockup_path or not self.design_path:
            print("Please select both mockup and design")
            return
            
        try:
            # If we have a live scene with background, render the scene directly.
            # This guarantees parity (includes watermark, blend modes, opacity, etc.).
            if self.bg_item is not None:
                scene_rect = self.bg_item.sceneBoundingRect()
                width = int(round(scene_rect.width()))
                height = int(round(scene_rect.height()))
                if width > 0 and height > 0:
                    img = QImage(width, height, QImage.Format_ARGB32)
                    # Fill transparent; background image will cover anyway
                    img.fill(Qt.transparent)
                    painter = QPainter(img)
                    painter.setRenderHints(QPainter.Antialiasing | QPainter.SmoothPixmapTransform, True)
                    self.scene.render(painter, QRectF(0, 0, width, height), scene_rect)
                    painter.end()
                    preview_dir = Path("preview_temp")
                    preview_dir.mkdir(exist_ok=True)
                    preview_path = preview_dir / f"preview_{self.design_path.stem}.png"
                    img.save(str(preview_path))
                    print(f"Preview (scene snapshot) saved: {preview_path}")
                    self._open_image_window(preview_path)
                    return

            # Prefer the live item's transformed geometry for parity
            if self.overlay_item is not None:
                item_rect = self.overlay_item.boundingRect()
                scene_center = self.overlay_item.mapToScene(item_rect.center())
                center_x = int(round(scene_center.x()))
                center_y = int(round(scene_center.y()))
                t = self.overlay_item.transform()
                base_w = item_rect.width()
                base_h = item_rect.height()
                scale_x = float(np.hypot(t.m11(), t.m21()))
                scale_y = float(np.hypot(t.m12(), t.m22()))
                target_w = max(1, int(round(base_w * scale_x)))
                target_h = max(1, int(round(base_h * scale_y)))
                rotation = float(np.degrees(np.arctan2(t.m21(), t.m11())))
            else:
                x, y = self.x_spin.value(), self.y_spin.value()
                w, h = self.w_spin.value(), self.h_spin.value()
                if self.anchor_combo.currentText() == "center":
                    center_x = x
                    center_y = y
                else:
                    center_x = x + w // 2
                    center_y = y + h // 2
                target_w, target_h = w, h
                rotation = self.rot_spin.value()
            
            result = render_simple(
                mockup_path=self.mockup_path,
                design_path=self.design_path,
                center_x_px=center_x,
                center_y_px=center_y,
                target_width_px=target_w,
                target_height_px=target_h,
                rotation_deg=rotation,
                maintain_aspect=self.aspect_combo.currentText(),
                opacity=self.opacity_spin.value(),
                blend_mode=self.blend_combo.currentText()
            )
            
            if result:
                # Save preview image
                preview_dir = Path("preview_temp")
                preview_dir.mkdir(exist_ok=True)
                preview_path = preview_dir / f"preview_{self.design_path.stem}.png"
                result.save(preview_path)
                print(f"Preview generated successfully: {preview_path}")
                # Open a viewing window with the generated image
                try:
                    self._open_image_window(preview_path)
                except Exception as e:
                    print(f"Failed to open preview window: {e}")
            else:
                print("Failed to generate preview")
                
        except Exception as e:
            print(f"Error generating preview: {e}")

    def _open_image_window(self, image_path: Path) -> None:
        try:
            win = QWidget()
            win.setWindowTitle(f"Preview - {image_path.name}")
            layout = QVBoxLayout(win)
            scroll = QScrollArea(win)
            scroll.setWidgetResizable(True)
            img_label = QLabel()
            img_label.setAlignment(Qt.AlignCenter)
            pix = QPixmap(str(image_path))
            img_label.setPixmap(pix)
            scroll.setWidget(img_label)
            layout.addWidget(scroll)
            win.resize(min(1400, pix.width()+60), min(1000, pix.height()+80))
            win.show()
            # Keep a reference so it's not GC'd
            if not hasattr(self, "_preview_windows"):
                self._preview_windows = []
            self._preview_windows.append(win)
        except Exception as e:
            print(f"Preview window error: {e}")
    
    def on_stop_clicked(self):
        """Request to stop current export pipeline"""
        self.stop_requested = True
        print("Stop requested")

    def reset_all(self):
        """Reset toàn bộ UI về mặc định và xóa mọi lựa chọn/đối tượng"""
        # Stop any running flow flag
        self.stop_requested = True

        # Clear scene and graphics items
        try:
            if self.scene is not None:
                self.scene.clear()
        except Exception:
            pass
        self.bg_item = None
        self.overlay_item = None
        self.wm_item = None
        self.active_item = None
        self.overlay_original_size = None
        self.wm_original_size = None
        self.mockup_size = None

        # Clear selected files/labels
        self.mockup_path = None
        self.design_path = None
        self.watermark_path = None
        if hasattr(self, 'mockup_label'):
            self.mockup_label.setText("Mockup: (chưa chọn)")
        if hasattr(self, 'design_label'):
            self.design_label.setText("Design: (chưa chọn)")
        if hasattr(self, 'watermark_label'):
            self.watermark_label.setText("Watermark: (không bắt buộc)")

        # Reset transform controls
        self.anchor_combo.setCurrentText("center")
        self.aspect_combo.setCurrentText("contain")
        self.blend_combo.setCurrentText("normal")
        self.x_spin.setValue(0)
        self.y_spin.setValue(0)
        self.w_spin.setValue(800)
        self.h_spin.setValue(600)
        self.rot_spin.setValue(0.0)
        self.opacity_spin.setValue(1.0)

        # Reset locks (no global lock)
        if hasattr(self, 'lock_design_chk'):
            self.lock_design_chk.setChecked(False)
        if hasattr(self, 'lock_watermark_chk'):
            self.lock_watermark_chk.setChecked(False)

        # Reset IO fields
        self.out_dir_edit.setText("outputs")
        self.source_folder_chk.setChecked(True)
        self.source_json_chk.setChecked(False)
        self.update_source_mode()
        self.designs_dir_edit.clear()
        self.json_path_edit.clear()
        self.temp_dir_edit.setText("designs_temp")
        self.pattern_edit.setText("{name}_mockup.png")

        # Reset upload/excel
        self.upload_chk.setChecked(False)
        self.imgbb_key_edit.clear()
        self.export_excel_chk.setChecked(False)
        self.excel_template_edit.clear()
        self.excel_out_edit.clear()

        # Reset progress and status
        self.progress_bar.setRange(0, 1)
        self.progress_bar.setValue(0)
        if hasattr(self, 'status_label'):
            self.status_label.setText("")

        # Set active layer back to design by default
        self.set_active_layer("design")
        print("All settings reset to defaults")

    # -------- Excel (5-variant) helpers --------
    def _excel_open(self) -> None:
        if not getattr(self, 'export_excel_chk', None) or not self.export_excel_chk.isChecked():
            self._excel_enabled = False
            return
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            print(f"openpyxl not installed: {e}")
            self._excel_enabled = False
            return
        self._excel_enabled = True
        tpl = self.excel_template_edit.text().strip() if hasattr(self, 'excel_template_edit') else ""
        if tpl and os.path.exists(tpl):
            self._excel_wb = openpyxl.load_workbook(tpl)
        else:
            self._excel_wb = openpyxl.Workbook()
        self._excel_ws = self._excel_wb.active
        self._excel_row_idx = 5
        self._excel_file_idx = 1
        # Max data rows per file (split by block of M rows)
        self._excel_max_data_rows = 1950
        # detect columns on header row 4
        ws = self._excel_ws
        header_row = 4
        self._excel_prodname_col = None
        self._excel_goods_col = None
        self._excel_sku_col = None
        self._excel_img_cols = []
        for idx, cell in enumerate(ws[header_row], start=1):
            val = str(cell.value or "").strip()
            low = val.lower()
            if val == "t_1_Product Name":
                self._excel_prodname_col = idx
            elif val == "t_1_Contribution Goods":
                self._excel_goods_col = idx
            elif val == "t_1_Contribution SKU":
                self._excel_sku_col = idx
            elif "sku images url" in low:
                self._excel_img_cols.append(idx)
        if not all([self._excel_prodname_col, self._excel_goods_col, self._excel_sku_col]) or len(self._excel_img_cols) == 0:
            print("Excel template missing required columns")
            self._excel_enabled = False
            return
        # cache prototype rows from row 5 downward until hit empty row
        self._excel_proto_rows = []
        r = 5
        while True:
            row_vals = [c.value for c in ws[r]]
            if all(v in (None, "") for v in row_vals):
                break
            row_map = {}
            for idx, cell in enumerate(ws[r], start=1):
                val = cell.value
                if val not in (None, ""):
                    row_map[idx] = val
            self._excel_proto_rows.append(row_map)
            r += 1
            if r - 5 > 200:  # safety guard
                break
        if len(self._excel_proto_rows) == 0:
            self._excel_proto_rows = [{} for _ in range(5)]
        # prepare output base (folder + filename)
        out_folder = self.excel_out_edit.text().strip() if hasattr(self, 'excel_out_edit') else ""
        if not out_folder:
            out_folder = self.out_dir_edit.text() or "outputs"
        Path(out_folder).mkdir(parents=True, exist_ok=True)
        self._excel_out_base = str(Path(out_folder) / "TEMU_results.xlsx")

    def _excel_save_current(self) -> None:
        if not self._excel_enabled or self._excel_wb is None:
            return
        base = Path(self._excel_out_base)
        out_path = base if self._excel_file_idx == 1 else base.with_stem(base.stem + f"_{self._excel_file_idx}")
        try:
            self._excel_wb.save(str(out_path))
            print(f"Saved Excel: {out_path}")
        except Exception as e:
            print(f"Failed to save Excel: {e}")

    def _excel_rotate(self) -> None:
        # save then open new workbook from template again
        self._excel_save_current()
        try:
            import openpyxl  # type: ignore
            tpl = self.excel_template_edit.text().strip() if hasattr(self, 'excel_template_edit') else ""
            if tpl and os.path.exists(tpl):
                self._excel_wb = openpyxl.load_workbook(tpl)
            else:
                self._excel_wb = openpyxl.Workbook()
            self._excel_ws = self._excel_wb.active
            self._excel_row_idx = 5
            self._excel_file_idx += 1
        except Exception as e:
            print(f"Excel rotate failed: {e}")
            self._excel_enabled = False

    def _excel_write_5(self, title: str, img_url: str) -> None:
        """Write M-variant block (M = number of prototype rows)."""
        if not self._excel_enabled or self._excel_ws is None:
            return
        M = len(getattr(self, '_excel_proto_rows', [])) or 5
        # If next block would exceed max rows, rotate before writing so split is a multiple of M
        try:
            max_rows = int(getattr(self, '_excel_max_data_rows', 1950))
        except Exception:
            max_rows = 1950
        data_rows_written = max(0, int(self._excel_row_idx) - 5)
        if data_rows_written + M > max_rows:
            self._excel_rotate()
        ws = self._excel_ws
        goods = "".join(random.choices(string.digits, k=9))
        for i in range(M):
            r = self._excel_row_idx + i
            # copy prototype defaults for each offset row
            proto = getattr(self, '_excel_proto_rows', [])
            row_map = proto[i] if len(proto) > i else {}
            for col_idx, val in row_map.items():
                ws.cell(row=r, column=col_idx, value=val)
            # override fields
            ws.cell(row=r, column=self._excel_prodname_col, value=title)
            ws.cell(row=r, column=self._excel_goods_col, value=goods)
            ws.cell(row=r, column=self._excel_sku_col, value=f"{goods}{i+1}")
            # fill only the FIRST SKU Images URL column per variant (as per rule)
            if self._excel_img_cols:
                ws.cell(row=r, column=self._excel_img_cols[0], value=img_url)
        self._excel_row_idx += M
    
    def on_export_batch(self):
        """Export batch of designs"""
        if not self.mockup_path:
            print("Please select a mockup first")
            return
            
        designs_pattern = self.designs_dir_edit.text()
        if not designs_pattern and not (hasattr(self, 'source_json_chk') and self.source_json_chk.isChecked()):
            print("Please specify design sources")
            return
            
        out_dir = self.out_dir_edit.text() or "outputs"
        pattern = self.pattern_edit.text()
        
        # Reset stop flag
        self.stop_requested = False

        # JSON streaming mode: process each item sequentially (download -> render -> upload)
        if hasattr(self, 'source_json_chk') and self.source_json_chk.isChecked():
            from PIL import Image  # noqa: F401
            from io import BytesIO
            import re as _re
            json_path = self.json_path_edit.text().strip() or self.designs_dir_edit.text().strip()
            if not json_path.lower().endswith('.json'):
                print("JSON mode selected but no JSON file provided")
                return
            temp_dir = Path(self.temp_dir_edit.text().strip() or "designs_temp").resolve()
            temp_dir.mkdir(parents=True, exist_ok=True)
            try:
                out_dir_res = Path(out_dir).resolve()
                if out_dir_res == temp_dir or str(out_dir_res).startswith(str(temp_dir) + os.sep):
                    temp_dir = temp_dir / "_json_cache"
                    temp_dir.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            try:
                items = self._resolve_json_items(json_path)
            except Exception as e:
                print(f"Error reading JSON: {e}")
                return
            def _slug(name: str) -> str:
                name = name.replace('_',' ').replace('-',' ')
                name = _re.sub(r"[^A-Za-z0-9 ]+","", name)
                name = _re.sub(r"\s+"," ", name).strip()
                return name or "image"
            total = len(items)
            self.progress_bar.setRange(0, max(1, total))
            self.progress_bar.setValue(0)
            uploaded_results: List[Dict[str, str]] = []
            # start excel session if enabled
            self._excel_open()
            for i, it in enumerate(items):
                if self.stop_requested:
                    break
                title = _slug(str(it.get('title','image')))
                url = str(it.get('url','')).strip()
                if not url:
                    self.progress_bar.setValue(i+1)
                    QApplication.processEvents()
                    continue
                # download if needed
                out_png = temp_dir / f"{title}.png"
                if not out_png.exists() or out_png.stat().st_size == 0:
                    try:
                        # Pre-check URL extension
                        path_part = urlsplit(url).path
                        ext = Path(path_part).suffix.lower()
                        if ext not in {'.png', '.jpg', '.jpeg', '.webp'}:
                            print(f"Skip non-image URL: {url}")
                            self.progress_bar.setValue(i+1)
                            QApplication.processEvents()
                            continue
                        resp = requests.get(url, timeout=30)
                        resp.raise_for_status()
                        ct = resp.headers.get('content-type', '')
                        if 'image' not in ct.lower():
                            print(f"Skip URL (not image content-type): {url}")
                            self.progress_bar.setValue(i+1)
                            QApplication.processEvents()
                            continue
                        buf = BytesIO(resp.content)
                        try:
                            img = Image.open(buf).convert("RGBA")
                        except Exception:
                            print(f"Skip URL (invalid image data): {url}")
                            self.progress_bar.setValue(i+1)
                            QApplication.processEvents()
                            continue
                        img.save(out_png)
                    except Exception as e:
                        print(f"Failed to fetch {title}: {e}")
                        self.progress_bar.setValue(i+1)
                        QApplication.processEvents()
                        continue
                # render one
                try:
                    # center/size/rotation from current overlay
                    w, h = self.w_spin.value(), self.h_spin.value()
                    if self.overlay_item is not None:
                        item_rect = self.overlay_item.boundingRect()
                        scene_center = self.overlay_item.mapToScene(item_rect.center())
                        center_x = int(round(scene_center.x()))
                        center_y = int(round(scene_center.y()))
                        t = self.overlay_item.transform()
                        base_w = item_rect.width()
                        base_h = item_rect.height()
                        scale_x = math.hypot(t.m11(), t.m21())
                        scale_y = math.hypot(t.m12(), t.m22())
                        target_w = max(1, int(round(base_w * scale_x)))
                        target_h = max(1, int(round(base_h * scale_y)))
                        rotation = math.degrees(math.atan2(t.m21(), t.m11()))
                    else:
                        x, y = self.x_spin.value(), self.y_spin.value()
                        if self.anchor_combo.currentText() == "center":
                            center_x, center_y = x, y
                        else:
                            center_x = x + w // 2
                            center_y = y + h // 2
                        target_w, target_h = w, h
                        rotation = self.rot_spin.value()
                    img_render = render_simple(
                        mockup_path=self.mockup_path,
                        design_path=out_png,
                        center_x_px=center_x,
                        center_y_px=center_y,
                        target_width_px=int(target_w),
                        target_height_px=int(target_h),
                        rotation_deg=float(rotation),
                        maintain_aspect=self.aspect_combo.currentText(),
                        opacity=self.opacity_spin.value(),
                        blend_mode=self.blend_combo.currentText()
                    )
                    if img_render:
                        filename = pattern.replace("{name}", Path(out_png).stem)
                        out_path = Path(out_dir) / filename
                        out_path.parent.mkdir(parents=True, exist_ok=True)
                        if out_path.suffix.lower() in {".jpg", ".jpeg"}:
                            img_render.convert("RGB").save(out_path, quality=95)
                        else:
                            img_render.save(out_path)
                        excel_img_url = out_path.as_posix()
                        if self.upload_chk.isChecked():
                            imgbb_key = self.imgbb_key_edit.text() or os.getenv('IMGBB_API_KEY')
                            if imgbb_key:
                                r = upload_to_imgbb(out_path, imgbb_key)
                                if r:
                                    uploaded_results.append({"title": title, "url": r.get("url",""), "id": r.get("id","")})
                                    excel_img_url = r.get("url", excel_img_url)
                        # write 5-variant block to excel (preserve required defaults)
                        self._excel_write_5(title, excel_img_url)
                except Exception as e:
                    print(f"Error rendering {title}: {e}")
                finally:
                    self.progress_bar.setValue(i+1)
                    QApplication.processEvents()
            # write uploads
            self._excel_save_current()
            if uploaded_results:
                try:
                    out_json = Path(out_dir) / "imgbb_results.json"
                    with open(out_json, "w", encoding="utf-8") as f:
                        json.dump(uploaded_results, f, ensure_ascii=False, indent=2)
                    print(f"Saved ImgBB results JSON: {out_json}")
                except Exception as e:
                    print(f"Failed to write results JSON: {e}")
            self.progress_bar.setRange(0,1)
            self.progress_bar.setValue(1 if not self.stop_requested else 0)
            return

        # Collect design paths
        design_paths = self._collect_design_paths(designs_pattern)
        if not design_paths:
            print("No design files found")
            return
            
        total = len(design_paths)
        print(f"Found {total} designs to process")
        # init progress
        self.progress_bar.setRange(0, total)
        self.progress_bar.setValue(0)
        
        # Process each design
        uploaded_results: List[Dict[str, str]] = []
        # Open Excel session for folder mode if enabled
        self._excel_open()
        for i, design_path in enumerate(design_paths):
            try:
                print(f"Processing {design_path.name} ({i+1}/{len(design_paths)})")
                if self.stop_requested:
                    break
                
                # Calculate center from the visual item to match preview exactly
                w, h = self.w_spin.value(), self.h_spin.value()
                if self.overlay_item is not None:
                    # Compute center from item's transformed center
                    item_rect = self.overlay_item.boundingRect()
                    scene_center = self.overlay_item.mapToScene(item_rect.center())
                    center_x = int(round(scene_center.x()))
                    center_y = int(round(scene_center.y()))
                    # Compute visual width/height from item's transform (scale applied)
                    t = self.overlay_item.transform()
                    base_w = item_rect.width()
                    base_h = item_rect.height()
                    # Extract scale from matrix (m11, m22)
                    scale_x = math.hypot(t.m11(), t.m21())
                    scale_y = math.hypot(t.m12(), t.m22())
                    target_w = max(1, int(round(base_w * scale_x)))
                    target_h = max(1, int(round(base_h * scale_y)))
                    # Extract rotation (degrees)
                    rotation = math.degrees(math.atan2(t.m21(), t.m11()))
                else:
                    # Fallback to anchor + spins
                    x, y = self.x_spin.value(), self.y_spin.value()
                    if self.anchor_combo.currentText() == "center":
                        center_x, center_y = x, y
                    else:
                        center_x = x + w // 2
                        center_y = y + h // 2
                    target_w, target_h = w, h
                    rotation = self.rot_spin.value()
                
                # Use original design size for export to ensure identical output size across designs
                # Export should match exactly the size user is seeing (spins w/h)
                target_w, target_h = w, h

                result = render_simple(
                    mockup_path=self.mockup_path,
                    design_path=design_path,
                    center_x_px=center_x,
                    center_y_px=center_y,
                    target_width_px=int(target_w),
                    target_height_px=int(target_h),
                    rotation_deg=float(rotation),
                    # Use 'contain' to keep the same strategy as preview sizing
                    maintain_aspect=self.aspect_combo.currentText(),
                    opacity=self.opacity_spin.value(),
                    blend_mode=self.blend_combo.currentText()
                )
                
                if result:
                    # Save the rendered image
                    from PIL import Image
                    output_name = pattern.replace("{name}", design_path.stem)
                    output_path = Path(out_dir) / output_name
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                    
                    if output_path.suffix.lower() in {".jpg", ".jpeg"}:
                        result_rgb = result.convert("RGB")
                        result_rgb.save(output_path, quality=95)
                    else:
                        result.save(output_path)
                    
                    print(f"Saved: {output_path}")
                    
                    # Default Excel image URL is local file path; replaced by ImgBB if uploaded
                    excel_img_url = output_path.as_posix()
                    # Upload to ImgBB if requested
                    if self.upload_chk.isChecked():
                        imgbb_key = self.imgbb_key_edit.text() or os.getenv('IMGBB_API_KEY')
                        if imgbb_key:
                            upload_result = upload_to_imgbb(output_path, imgbb_key)
                            if upload_result:
                                # Collect page URL (e.g. https://ibb.co/<code>)
                                title_for_result = design_path.stem
                                uploaded_results.append({
                                    "title": title_for_result,
                                    "url": upload_result.get("url", ""),
                                    "id": upload_result.get("id", "")
                                })
                                print(f"Uploaded to ImgBB: {upload_result.get('url','')}")
                                excel_img_url = upload_result.get("url", excel_img_url)
                            else:
                                print("Failed to upload to ImgBB")
                    # Write 5-variant block to Excel if enabled
                    self._excel_write_5(design_path.stem, excel_img_url)
                else:
                    print(f"Failed to render {design_path.name}")
                
            except Exception as e:
                print(f"Error processing {design_path.name}: {e}")
            finally:
                self.progress_bar.setValue(i+1)
        
        # Save Excel at the end for folder mode
        self._excel_save_current()
        # If uploaded, write a JSON summary mapping title -> page URL
        if uploaded_results:
            try:
                out_json = Path(out_dir) / "imgbb_results.json"
                with open(out_json, "w", encoding="utf-8") as f:
                    json.dump(uploaded_results, f, ensure_ascii=False, indent=2)
                print(f"Saved ImgBB results JSON: {out_json}")
                print(json.dumps(uploaded_results, ensure_ascii=False))
            except Exception as e:
                print(f"Failed to write results JSON: {e}")
        self.status_label.setText("Hoàn thành")
        self.progress_bar.setRange(0, 1)
        self.progress_bar.setValue(1)
        QMessageBox.information(self, "Hoàn thành", "Export hàng loạt đã hoàn thành.")
    
    def _collect_design_paths(self, pattern: str) -> List[Path]:
        """Collect design file paths from pattern"""
        paths = []
        
        if pattern.endswith('.json'):
            # JSON file with URLs
            try:
                with open(pattern, 'r') as f:
                    data = json.load(f)
                urls = data.get('urls', [])
                # Convert URLs to temp file paths
                temp_dir = Path("designs_temp")
                if temp_dir.exists():
                    paths = list(temp_dir.glob("*"))
            except Exception as e:
                print(f"Error reading JSON: {e}")
        else:
            # File pattern or directory
            if os.path.isdir(pattern):
                # Directory path
                paths = list(Path(pattern).glob("*.png"))
                paths.extend(Path(pattern).glob("*.jpg"))
                paths.extend(Path(pattern).glob("*.jpeg"))
            elif os.path.isfile(pattern):
                # Single file
                paths = [Path(pattern)]
            else:
                # Glob pattern
                try:
                    # Use glob.glob for absolute paths
                    if os.path.isabs(pattern):
                        glob_paths = glob.glob(pattern)
                    else:
                        glob_paths = glob.glob(pattern)
                    paths = [Path(p) for p in glob_paths]
                except Exception as e:
                    print(f"Error with glob pattern: {e}")
        
        return paths
    
    def eventFilter(self, obj, event):
        """Handle viewport events for interactions"""
        if obj == self.view.viewport():
            if event.type() == QEvent.Wheel:
                # Bên trong khung preview: wheel => phóng/thu (Shift để xoay)
                if not self.active_item:
                    return False
                if (self.active_item == self.overlay_item and self.design_locked) or \
                   (self.active_item == self.wm_item and self.watermark_locked):
                    return False
                return self.handle_wheel_event(event)
            elif event.type() == QEvent.MouseButtonPress:
                return self.handle_mouse_press(event)
            elif event.type() == QEvent.MouseButtonDblClick:
                return self.handle_double_click(event)
        
        return super().eventFilter(obj, event)
    
    def handle_wheel_event(self, event: QWheelEvent):
        """Handle wheel events for scaling and rotation"""
        if not self.active_item:
            return False
            
        # Check if item is locked
        # Wheel zoom must still work even if the item is locked
        
        delta = event.angleDelta().y()
        
        if event.modifiers() & Qt.ShiftModifier:
            # Shift + wheel = rotate
            rotation_change = delta / 120.0 * 5  # 5 degrees per wheel step
            current_rotation = self.rot_spin.value()
            self.rot_spin.setValue(current_rotation + rotation_change)
        else:
            # Wheel = scale
            # Keep visual center fixed while scaling
            try:
                rect_local = self.active_item.boundingRect()
                center_scene = self.active_item.mapToScene(rect_local.center())
            except Exception:
                center_scene = QPointF(float(self.x_spin.value()), float(self.y_spin.value()))
            # Mark scaling mode so update respects center anchor
            self._is_scaling = True
            # Ensure X/Y represent the current visual center
            self.x_spin.setValue(int(round(center_scene.x())))
            self.y_spin.setValue(int(round(center_scene.y())))
            scale_factor = 1.1 if delta > 0 else 0.9
            current_w = max(1, self.w_spin.value())
            current_h = max(1, self.h_spin.value())
            self.w_spin.setValue(max(1, int(current_w * scale_factor)))
            self.h_spin.setValue(max(1, int(current_h * scale_factor)))
            self._is_scaling = False
        
        return True
    
    def handle_mouse_press(self, event: QMouseEvent):
        """Handle mouse press events"""
        # no global lock feature
            
        # Check if clicking on an item to select it
        pos = self.view.mapToScene(event.pos())
        items = self.scene.items(pos)
        
        for item in items:
            if isinstance(item, BlendPixmapItem):
                # Chỉ đổi layer active; KHÔNG cập nhật spin khi layer đang khóa
                if item == self.overlay_item:
                    self.set_active_layer("design")
                elif item == self.wm_item:
                    self.set_active_layer("watermark")
                break
        return False
    
    def handle_double_click(self, event: QMouseEvent):
        """Handle double-click events"""
        if not self.active_item:
            return False
            
        # If the active layer is locked, ignore double-click centering
        if (self.active_item == self.overlay_item and self.design_locked) or \
           (self.active_item == self.wm_item and self.watermark_locked):
            return True

        # Double-click to center the active item
        self.center_overlay()
        return True
    
    def keyPressEvent(self, event: QKeyEvent):
        """Handle key press events"""
        if event.key() == Qt.Key_D:
            # D key to focus design
            self.set_active_layer("design")
        elif event.key() == Qt.Key_W:
            # W key to focus watermark
            self.set_active_layer("watermark")
        elif event.key() == Qt.Key_C:
            # C key to center active item
            self.center_overlay()
        elif event.key() == Qt.Key_R:
            # R key to reset active item
            self.reset_overlay()
        else:
            super().keyPressEvent(event)

def main():
    app = QApplication(sys.argv)
    
    # Set application style
    app.setStyle("Fusion")
    # Load external QSS theme if present
    try:
        qss_path = Path(__file__).resolve().parent / "styles.qss"
        if qss_path.exists():
            with open(qss_path, "r", encoding="utf-8") as f:
                app.setStyleSheet(f.read())
    except Exception as e:
        print(f"Failed to load styles.qss: {e}")
    
    # Create and show main window
    window = MockupApp()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()


